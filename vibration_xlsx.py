#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Aug 13 17:33:41 2018

@author: chenjian
"""

import re,os,sys,codecs
from openpyxl import load_workbook
from collections import Counter

#list of all files in dir
if len(sys.argv)>1:
    
	arr = os.listdir(sys.argv[1])
else:
	arr = os.listdir()

#print (arr)

arr_excel = [b for b in arr if b[-5:]=='.xlsx']
arr_excel = [b for b in arr_excel if b!='charging_port_2018.xlsx']
print (arr_excel)

#%%

def get_list_of_vib_sents(ws_input):
    
    vib_logs = []

    list_logs = []
    
    for i in range(2,ws_input.max_row+1):
        
        chatid = ws_input['A'+str(i)].value
        
        chatlog = ws_input['Q'+str(i)].value
        
        if ws_input['D'+str(i)].value is not None:
            name_agent = ws_input['D'+str(i)].value[3:8]
        
        #print (name_agent)
        chatturns = re.split('\n',chatlog)
    
        for index,chatturn in enumerate(chatturns):
            
            if name_agent in chatturn:
                continue
             
            if 'charger' in chatturn:
                flag_skip = 0
                
 
                vib_logs.append(chatturn)
                
                item = chatturn
                
                list_words = item.split()
                
                for n_word in ['replace','replacement','buy','bought','you','your','we','wireless','different','other','multiple']:
                    
                    if n_word in list_words:
                        
                        flag_skip = 1
                
                if flag_skip == 1:
                    continue
                
                sent = item[item.find(':',10)+2:]
                print (str(chatid)+"-"+str(index),sent,len(sent.split()))
                list_logs.append((str(chatid)+"-"+str(index),sent,len(sent.split())))
                
                continue
        """
                for ind, word in enumerate(list_words):
                    
                    if 'battery' in word:
                        
                        for n_word in ['enlarge','swell','']:
                            
                            if n_word in list_words[ind-5:ind+5]:
                                
                                sent = item[item.find(':',10)+2:]
        """                        
                                
                                

    return list_logs
            
#%%

wb_output = load_workbook(filename = '/Users/chenjian/Lenovo/20180704_preExtractCase_new/chatturn_temp.xlsx')
ws_output = wb_output.worksheets[0]

output_row = 2
#%%
for item in arr_excel:
    
    print (item)
    
    wb_input = load_workbook(filename = item)
    
    ws_input = wb_input.worksheets[0]

    output_list = set(get_list_of_vib_sents(ws_input))
    
    for tuple_output in output_list:
        
        if tuple_output[2]<5:
            continue
        ws_output['A'+str(output_row)].value = tuple_output[0]
        ws_output['B'+str(output_row)].value = tuple_output[1]
        ws_output['C'+str(output_row)].value = tuple_output[2]
        print (output_row, tuple_output)
        output_row +=1
        
#%%

wb_output.save(filename='/Users/chenjian/Lenovo/0626/charging_port/charger_excel.xlsx')