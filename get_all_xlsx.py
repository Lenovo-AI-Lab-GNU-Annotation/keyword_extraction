#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Oct 15 20:35:43 2018

@author: chenjian
"""

import re
import csv
import sys
import os
from openpyxl import load_workbook
from collections import Counter

#%%

list_output = [['chatturn_id','content','token_number']]


#%%

#list of all files in dir
if len(sys.argv)>1:
    
	arr = os.listdir(sys.argv[1])
else:
	arr = os.listdir()

#print (arr)

arr_excel = [b for b in arr if b[-5:]=='.xlsx']
print (arr_excel)

#%%



#%%

def get_list_of_vib_sents(ws_input):


    list_logs = []

    for i in range(2,ws_input.max_row+1):
        
        chatid = ws_input['N'+str(i)].value
        
        chatlog = ws_input['Q'+str(i)].value
        
        if ws_input['D'+str(i)].value is not None:
            name_agent = ws_input['D'+str(i)].value[3:8]
            
        #print (name_agent)
        chatturns = re.split('\n',chatlog)
        
        for index,chatturn in enumerate(chatturns):
            
            if name_agent in chatturn:
                continue
            
            item = chatturn
            
            sent = item[item.find(':',10)+2:]
                        
            list_logs.append((str(chatid)+"-"+str(index),sent,len(sent.split())))

            continue                               

    return list_logs
            

#%%


for item in arr_excel:
        
    wb_input = load_workbook(filename = item)
    
    ws_input = wb_input.worksheets[0]

    output_list = set(get_list_of_vib_sents(ws_input))
    
    for tuple_output in output_list:
        
        if tuple_output[2]<5:
            continue
        
#        ws_output['A'+str(output_row)].value = tuple_output[0]
#        ws_output['B'+str(output_row)].value = tuple_output[1]
#        ws_output['C'+str(output_row)].value = tuple_output[2]
        
        list_output.append([tuple_output[0],tuple_output[1],tuple_output[2]])
        
        print (len(list_output), tuple_output)

#%%

with open('all_xlsx.csv', 'w') as csvFile:
    writer = csv.writer(csvFile)
    writer.writerows(list_output)
csvFile.close()